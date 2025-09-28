import os
import json
import logging
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config import Config

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ReportGenerator:
    def __init__(self):
        self.config = Config()
        self.output_dir = self.config.OUTPUT_DIR
        self._ensure_output_dir()
        
    def _ensure_output_dir(self):
        """Ensure output directory exists"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"Created output directory: {self.output_dir}")
    
    def generate_comprehensive_report(self, jobs: List[Dict], contacts_summary: Dict) -> str:
        """Generate a comprehensive report with all job details and HR contacts"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Generate multiple report formats
        json_report = self._generate_json_report(jobs, contacts_summary, timestamp)
        excel_report = self._generate_excel_report(jobs, contacts_summary, timestamp)
        html_report = self._generate_html_report(jobs, contacts_summary, timestamp)
        
        logger.info(f"Generated reports: {json_report}, {excel_report}, {html_report}")
        return excel_report  # Return Excel report as primary
    
    def _generate_json_report(self, jobs: List[Dict], contacts_summary: Dict, timestamp: str) -> str:
        """Generate JSON report"""
        report_data = {
            'generated_at': datetime.now().isoformat(),
            'summary': contacts_summary,
            'jobs': jobs,
            'statistics': self._calculate_statistics(jobs)
        }
        
        filename = f"job_report_{timestamp}.json"
        filepath = os.path.join(self.output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
        
        return filepath
    
    def _generate_excel_report(self, jobs: List[Dict], contacts_summary: Dict, timestamp: str) -> str:
        """Generate Excel report with multiple sheets"""
        filename = f"job_report_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        self._create_summary_sheet(wb, contacts_summary, jobs)
        
        # Create jobs sheet
        self._create_jobs_sheet(wb, jobs)
        
        # Create contacts sheet
        self._create_contacts_sheet(wb, jobs)
        
        # Create company analysis sheet
        self._create_company_analysis_sheet(wb, jobs)
        
        wb.save(filepath)
        return filepath
    
    def _create_summary_sheet(self, wb: Workbook, contacts_summary: Dict, jobs: List[Dict]):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Job Search Report Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Report metadata
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1
        
        ws[f'A{row}'] = "Search Location:"
        ws[f'B{row}'] = self.config.LOCATION
        row += 1
        
        ws[f'A{row}'] = "Experience Level:"
        ws[f'B{row}'] = self.config.EXPERIENCE_LEVEL
        row += 2
        
        # Key statistics
        ws[f'A{row}'] = "KEY STATISTICS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        stats = self._calculate_statistics(jobs)
        
        ws[f'A{row}'] = "Total Jobs Found:"
        ws[f'B{row}'] = stats.get('total_jobs', 0)
        row += 1
        
        ws[f'A{row}'] = "Jobs with HR Contacts:"
        ws[f'B{row}'] = stats.get('jobs_with_contacts', 0)
        row += 1
        
        ws[f'A{row}'] = "Total HR Contacts:"
        ws[f'B{row}'] = stats.get('total_contacts', 0)
        row += 1
        
        ws[f'A{row}'] = "Unique Companies:"
        ws[f'B{row}'] = stats.get('unique_companies', 0)
        row += 1
        
        ws[f'A{row}'] = "Average Relevance Score:"
        ws[f'B{row}'] = round(stats.get('average_relevance_score', 0), 2)
        row += 2
        
        # Top companies
        ws[f'A{row}'] = "TOP COMPANIES"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        top_companies = stats.get('top_companies', {})
        for company, count in list(top_companies.items())[:10]:
            ws[f'A{row}'] = company
            ws[f'B{row}'] = count
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_jobs_sheet(self, wb: Workbook, jobs: List[Dict]):
        """Create detailed jobs sheet"""
        ws = wb.create_sheet("Job Details")
        
        # Headers
        headers = [
            'Company', 'Job Title', 'Location', 'Source', 'Posted Date', 
            'Relevance Score', 'HR Contacts Count', 'Company Website',
            'Job URL', 'Description Preview'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row_idx, job in enumerate(jobs, 2):
            ws.cell(row=row_idx, column=1, value=job.get('company', ''))
            ws.cell(row=row_idx, column=2, value=job.get('title', ''))
            ws.cell(row=row_idx, column=3, value=job.get('location', ''))
            ws.cell(row=row_idx, column=4, value=job.get('source', ''))
            ws.cell(row=row_idx, column=5, value=job.get('posted_date', ''))
            ws.cell(row=row_idx, column=6, value=job.get('relevance_score', 0))
            ws.cell(row=row_idx, column=7, value=len(job.get('hr_contacts', [])))
            
            # Company info
            company_info = job.get('company_info', {})
            ws.cell(row=row_idx, column=8, value=company_info.get('website', ''))
            ws.cell(row=row_idx, column=9, value=job.get('url', ''))
            
            # Description preview (first 200 chars)
            description = job.get('description', '')
            ws.cell(row=row_idx, column=10, value=description[:200] + '...' if len(description) > 200 else description)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_contacts_sheet(self, wb: Workbook, jobs: List[Dict]):
        """Create HR contacts sheet"""
        ws = wb.create_sheet("HR Contacts")
        
        # Headers
        headers = [
            'Company', 'Job Title', 'Contact Name', 'Contact Title', 
            'Email', 'Phone', 'LinkedIn URL'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        row_idx = 2
        for job in jobs:
            contacts = job.get('hr_contacts', [])
            for contact in contacts:
                ws.cell(row=row_idx, column=1, value=job.get('company', ''))
                ws.cell(row=row_idx, column=2, value=job.get('title', ''))
                ws.cell(row=row_idx, column=3, value=contact.get('name', ''))
                ws.cell(row=row_idx, column=4, value=contact.get('title', ''))
                ws.cell(row=row_idx, column=5, value=contact.get('email', ''))
                ws.cell(row=row_idx, column=6, value=contact.get('phone', ''))
                ws.cell(row=row_idx, column=7, value=contact.get('linkedin_url', ''))
                row_idx += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_company_analysis_sheet(self, wb: Workbook, jobs: List[Dict]):
        """Create company analysis sheet"""
        ws = wb.create_sheet("Company Analysis")
        
        # Group jobs by company
        company_data = {}
        for job in jobs:
            company = job.get('company', '')
            if company not in company_data:
                company_data[company] = {
                    'jobs': [],
                    'contacts': [],
                    'company_info': job.get('company_info', {})
                }
            company_data[company]['jobs'].append(job)
            company_data[company]['contacts'].extend(job.get('hr_contacts', []))
        
        # Headers
        headers = [
            'Company', 'Job Count', 'HR Contacts', 'Website', 'Industry',
            'Employee Count', 'Description', 'Job Titles'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row_idx, (company, data) in enumerate(company_data.items(), 2):
            company_info = data['company_info']
            job_titles = [job.get('title', '') for job in data['jobs']]
            
            ws.cell(row=row_idx, column=1, value=company)
            ws.cell(row=row_idx, column=2, value=len(data['jobs']))
            ws.cell(row=row_idx, column=3, value=len(data['contacts']))
            ws.cell(row=row_idx, column=4, value=company_info.get('website', ''))
            ws.cell(row=row_idx, column=5, value=company_info.get('industry', ''))
            ws.cell(row=row_idx, column=6, value=company_info.get('employee_count', ''))
            ws.cell(row=row_idx, column=7, value=company_info.get('description', ''))
            ws.cell(row=row_idx, column=8, value=', '.join(job_titles))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _generate_html_report(self, jobs: List[Dict], contacts_summary: Dict, timestamp: str) -> str:
        """Generate HTML report"""
        filename = f"job_report_{timestamp}.html"
        filepath = os.path.join(self.output_dir, filename)
        
        stats = self._calculate_statistics(jobs)
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Job Search Report - {datetime.now().strftime('%Y-%m-%d')}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ background-color: #f0f0f0; padding: 20px; border-radius: 5px; }}
                .stats {{ display: flex; justify-content: space-around; margin: 20px 0; }}
                .stat-box {{ background-color: #e8f4f8; padding: 15px; border-radius: 5px; text-align: center; }}
                .job-card {{ border: 1px solid #ddd; margin: 10px 0; padding: 15px; border-radius: 5px; }}
                .contact-info {{ background-color: #f9f9f9; padding: 10px; margin: 5px 0; border-radius: 3px; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Job Search Report</h1>
                <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>Location: {self.config.LOCATION} | Experience Level: {self.config.EXPERIENCE_LEVEL}</p>
            </div>
            
            <div class="stats">
                <div class="stat-box">
                    <h3>{stats.get('total_jobs', 0)}</h3>
                    <p>Total Jobs</p>
                </div>
                <div class="stat-box">
                    <h3>{stats.get('jobs_with_contacts', 0)}</h3>
                    <p>Jobs with Contacts</p>
                </div>
                <div class="stat-box">
                    <h3>{stats.get('total_contacts', 0)}</h3>
                    <p>HR Contacts</p>
                </div>
                <div class="stat-box">
                    <h3>{stats.get('unique_companies', 0)}</h3>
                    <p>Unique Companies</p>
                </div>
            </div>
            
            <h2>Job Details</h2>
        """
        
        for job in jobs:
            contacts = job.get('hr_contacts', [])
            company_info = job.get('company_info', {})
            
            html_content += f"""
            <div class="job-card">
                <h3>{job.get('title', '')} at {job.get('company', '')}</h3>
                <p><strong>Location:</strong> {job.get('location', '')}</p>
                <p><strong>Source:</strong> {job.get('source', '')}</p>
                <p><strong>Posted:</strong> {job.get('posted_date', '')}</p>
                <p><strong>Relevance Score:</strong> {job.get('relevance_score', 0)}</p>
                <p><strong>Company Website:</strong> <a href="{company_info.get('website', '')}" target="_blank">{company_info.get('website', '')}</a></p>
                <p><strong>Job URL:</strong> <a href="{job.get('url', '')}" target="_blank">View Job</a></p>
                
                <h4>HR Contacts ({len(contacts)})</h4>
            """
            
            for contact in contacts:
                html_content += f"""
                <div class="contact-info">
                    <p><strong>{contact.get('name', '')}</strong> - {contact.get('title', '')}</p>
                    <p>Email: {contact.get('email', '')}</p>
                    <p>Phone: {contact.get('phone', '')}</p>
                    <p>LinkedIn: <a href="{contact.get('linkedin_url', '')}" target="_blank">Profile</a></p>
                </div>
                """
            
            html_content += "</div>"
        
        html_content += """
        </body>
        </html>
        """
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return filepath
    
    def _calculate_statistics(self, jobs: List[Dict]) -> Dict:
        """Calculate statistics for the jobs"""
        if not jobs:
            return {}
        
        total_jobs = len(jobs)
        jobs_with_contacts = len([job for job in jobs if job.get('hr_contacts')])
        total_contacts = sum(len(job.get('hr_contacts', [])) for job in jobs)
        
        # Unique companies
        companies = set(job.get('company', '') for job in jobs if job.get('company'))
        unique_companies = len(companies)
        
        # Average relevance score
        relevance_scores = [job.get('relevance_score', 0) for job in jobs]
        avg_relevance = sum(relevance_scores) / len(relevance_scores) if relevance_scores else 0
        
        # Sources
        sources = {}
        for job in jobs:
            source = job.get('source', 'Unknown')
            sources[source] = sources.get(source, 0) + 1
        
        # Top companies
        company_counts = {}
        for job in jobs:
            company = job.get('company', '')
            if company:
                company_counts[company] = company_counts.get(company, 0) + 1
        
        top_companies = dict(sorted(company_counts.items(), key=lambda x: x[1], reverse=True)[:10])
        
        return {
            'total_jobs': total_jobs,
            'jobs_with_contacts': jobs_with_contacts,
            'total_contacts': total_contacts,
            'unique_companies': unique_companies,
            'average_relevance_score': avg_relevance,
            'sources': sources,
            'top_companies': top_companies
        }

if __name__ == "__main__":
    # Test the report generator
    generator = ReportGenerator()
    
    # Sample test data
    test_jobs = [
        {
            'title': 'Software Engineer - Fresher',
            'company': 'Tech Corp',
            'location': 'Bangalore',
            'source': 'LinkedIn',
            'posted_date': '2024-01-15',
            'relevance_score': 8,
            'hr_contacts': [
                {
                    'name': 'John Doe',
                    'title': 'HR Manager',
                    'email': 'john@techcorp.com',
                    'phone': '+91-9876543210',
                    'linkedin_url': 'https://linkedin.com/in/johndoe'
                }
            ],
            'company_info': {
                'website': 'https://techcorp.com',
                'industry': 'Technology',
                'employee_count': '100-500'
            }
        }
    ]
    
    contacts_summary = {
        'total_jobs': 1,
        'jobs_with_contacts': 1,
        'total_contacts': 1
    }
    
    report_path = generator.generate_comprehensive_report(test_jobs, contacts_summary)
    print(f"Generated report: {report_path}")
